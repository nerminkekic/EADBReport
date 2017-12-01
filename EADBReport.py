"""
This script will obtain four peaces of information for all ASP sites.
    1. Total number of Archives
    2. Total number or Retrieves
    3. Total number of Archives in GB
    4. Total number of Retrieves in GB
"""
import pyodbc
import datetime
import smtplib
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders


def eadbreport():

    """
    This function will do 4 tasks:
    1.  Connect to SQL Server.
    2.  Run query and retrieve data from SQL server Data Base.
    3.  Write the data to Excel worksheet.
    4.  Email the worksheet to end user.
    """

    # File name for Excel Worksheet
    excel_filename = ""

    # Define dictionary to store data for processing
    eaData = {'sep':[9, 0, 0, 0, 0], 'oct': [10, 0, 0, 0, 0], 'nov': [11, 0, 0, 0, 0]}

    # Get all Virtual Archive names from SQL Server
    virtual_archives = get_archives()

    # Set up Excel Worksheet.
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "EADBReport"
    work_sheet.append(['Month', 'Archives', 'GB Archives', 'Retrieves', 'GB Retrieves'])

    # Assign font and background color properties for Column Title cells
    f = Font(name="Arial", size=14, bold=True, color="FF000000")
    fill = PatternFill(fill_type="solid", start_color="00FFFF00")

    # Format Worksheet columns
    for L in "ABCDE":
        work_sheet[L + "1"].fill = fill
        work_sheet[L + "1"].font = f
        work_sheet.column_dimensions[L].width = 35.0

    # Obtain Exam Volume for all virtual archives and write the data to excel sheet.
    for archive in virtual_archives:
        # Obtains archive volume from SQL server.
        rows = archive_volume(archive)
        for row in rows:
            if row[0] == 9:
                eaData['sep'][1] += row[1]      # Sum Total Archive
                eaData['sep'][2] += row[2]      # Sum Total Archives GB
            if row[0] == 10:
                eaData['oct'][1] += row[1]      # Sum Total Archive
                eaData['oct'][2] += row[2]      # Sum Total Archives GB
            if row[0] == 11:
                eaData['nov'][1] += row[1]      # Sum Total Archive
                eaData['nov'][2] += row[2]      # Sum Total Archives GB

        # Sum Retrieve volume and GB
        rows = get_retrieves(archive)
        for row in rows:
            if row[0] == 9:
                eaData['sep'][3] += row[1]                      # Sum Total Archive
                eaData['sep'][4] += convert_to_gb(row[2])       # Sum Total Archives GB
            if row[0] == 10:
                eaData['oct'][3] += row[1]                      # Sum Total Archive
                eaData['oct'][4] += convert_to_gb(row[2])       # Sum Total Archives GB
            if row[0] == 11:
                eaData['nov'][3] += row[1]                      # Sum Total Archive
                eaData['nov'][4] += convert_to_gb(row[2])       # Sum Total Archives GB

        print("Added {} Data to Workbook!".format(archive))

    # Adds sum of archives, retrieves, and GB per month
    for key in eaData.keys():
        work_sheet.append([eaData[key][0],              # Month
                           eaData[key][1],              # Sum of Archives
                           eaData[key][2],              # Sum of Archives in GB
                           eaData[key][3],              # Sum of Retrieves
                           eaData[key][4],              # Sum of Retrieves in GB
                           ])

    # Format cells to use 1000 comma separator.
    work_sheet['C{}'.format(work_sheet.max_row)].style = 'Comma [0]'
    work_sheet['D{}'.format(work_sheet.max_row)].style = 'Comma'
    work_sheet['E{}'.format(work_sheet.max_row)].style = 'Comma'
    work_sheet['F{}'.format(work_sheet.max_row)].style = 'Comma'
    work_sheet['E{}'.format(work_sheet.max_row)].style = 'Comma'

    # Saves Excel worksheet.
    excel_filename = "EADBReport_{}.xlsx".format(datetime.datetime.now().strftime("%Y-%m-%d"))
    work_book.save(excel_filename)

    # Send email with attachment.
    # send_email(excel_filename)
    #


# Obtain Archive Names from SQL Server.
def get_archives():
    """
    This function will return list of Virtual Archive Names from SQL Server.
    """
    # Create list to store archive names.
    archive_list = []

    # Read file for credentials
    with open("data.txt", "r") as f:
        read_data = f.readline().split()

    # Define data base connection parameters.
    sqlserver = read_data[2]
    database = 'RSAdmin'
    username = read_data[0]
    password = read_data[1]

    # Establish DB connections.
    conn = pyodbc.connect(
        r'DRIVER={SQL Server};'
        r'SERVER=' + sqlserver + ';'
        r'DATABASE=' + database + ';'
        r'UID=' + username + ';'
        r'PWD=' + password + ''
    )
    cur = conn.cursor()
    # Execute query on Data Base.
    cur.execute("""
                SELECT DBName from tblArchive
                ORDER BY DBName
                """)
    rows = cur.fetchall()

    # Add Archive names to archive list.
    for row in rows:
        archive_list.append(row[0])
    # Close SQL Connection.
    cur.close()
    conn.close()

    return archive_list


# Obtain Retrieves volume from SQL Server.
def get_retrieves(db_name):
    """
    This function will obtain retrieve volume form SQL server.
    """

    # Read file for credentials
    with open("data.txt", "r") as f:
        read_data = f.readline().split()

    # Define data base connection parameters.
    sqlserver = read_data[2]
    database = db_name
    username = read_data[0]
    password = read_data[1]

    # Establish DB connections.
    conn = pyodbc.connect(
        r'DRIVER={SQL Server};'
        r'SERVER=' + sqlserver + ';'
        r'DATABASE=' + database + ';'
        r'UID=' + username + ';'
        r'PWD=' + password + ''
    )
    cur = conn.cursor()
    # Execute query on Data Base.
    cur.execute("""
                select month(DayStart) as retreiveDay, count(StudyUID) as retreiveVolume, sum(BytesSent/1024/1024) as sumofMB from tblAuditTrailDicom with (nolock)
                where command = 16385 and completioncode =0and daystart >'20170901 07:30:00.000' and daystart <'20171129 07:30:00.000'
                group by month(DayStart)
                order by month(DayStart)  
                """)
    rows = cur.fetchall()
    # Close SQL Server Connections.
    cur.close()
    conn.close()
    return rows

# Obtain archive volume.
def archive_volume(db_name):
    """
    This function will obtain archive volume form SQL server.
    """

    # Read file for credentials
    with open("data.txt", "r") as f:
        read_data = f.readline().split()

    # Define data base connection parameters.
    sqlserver = read_data[2]
    database = db_name
    username = read_data[0]
    password = read_data[1]

    # Establish DB connections.
    conn = pyodbc.connect(
        r'DRIVER={SQL Server};'
        r'SERVER='+sqlserver+';'
        r'DATABASE='+database+';'
        r'UID='+username+';'
        r'PWD='+password+''
        )
    cur = conn.cursor()
    # Execute query on Data Base.
    cur.execute("""
                set transaction isolation level read uncommitted
                select month(firstarchivedate) as studymonth,
                count(distinct id1) as StudyCount, sum(bytesize/1024/1024/1024) as sumofGB
                from ((tbldicomstudy left join tbldicomseries on tbldicomstudy.id1=tbldicomseries._id1)left join  tblfile on tbldicomseries.id2=tblfile._id2file)
                where firstarchivedate > '2017-09-01' and firstarchivedate <'2017-11-30'
                group by  month(firstarchivedate)
                order by  Month(firstarchivedate)
                """)
    rows = cur.fetchall()
    # Close SQL Server Connections.
    cur.close()
    conn.close()
    return rows


# Send email with Report
def send_email(file_attachment):
    """This function will send email with the attachment.
    It takes attachment file name as argument.
    """

    # Define email body
    body = "This is EA Monthly report. See attached file for Total Exam Volume for each customer."
    content = MIMEText(body, 'plain')

    # Open file attachment
    filename = file_attachment
    infile = open(filename, "rb")

    # Set up attachment to be send in email
    part = MIMEBase("application", "octet-stream")
    part.set_payload(infile.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)

    msg = MIMEMultipart("alternative")

    # Define email recipients
    to_email = ["na@na.com"
        ]
    # Define From email
    from_email = "na@na.com"

    # Create email content
    msg["Subject"] = "ASP Monthly Report {}".format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    msg["From"] = from_email
    msg["To"] = ",".format(to_email)
    msg.attach(part)
    msg.attach(content)
    # Send email to SMTP server
    s = smtplib.SMTP("10.4.1.1", 25)
    s.sendmail(from_email, to_email, msg.as_string())
    s.close()


# Convert from MB to GB
def convert_to_gb(size_in_mb):
    """Convert Average exam size in MB to GB"""
    return round((size_in_mb / 1024), 2)


# Run script
eadbreport()
